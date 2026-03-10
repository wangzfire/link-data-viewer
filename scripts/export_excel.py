#!/usr/bin/env python3
"""
TSV → Excel 导出脚本
用法: python3 export_excel.py <输入TSV文件> <输出XLSX文件>

将 mysql -B 输出的 TSV 格式数据转换为格式化的 Excel 文件。
功能：
  - 自动列宽调整
  - 表头加粗 + 背景色
  - 数据行交替底色
  - 自动冻结首行（表头）
  - 自动启用筛选器
  - NULL 值标灰显示
"""

import sys
import os

def main():
    if len(sys.argv) != 3:
        print(f"用法: {sys.argv[0]} <输入TSV文件> <输出XLSX文件>")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    if not os.path.exists(input_path):
        print(f"[错误] 输入文件不存在: {input_path}")
        sys.exit(1)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[错误] openpyxl 模块未安装，请执行: pip install openpyxl")
        sys.exit(1)

    # 读取 TSV 数据
    rows = []
    with open(input_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n\r")
            rows.append(line.split("\t"))

    if not rows:
        print("[错误] 输入文件为空")
        sys.exit(1)

    headers = rows[0]
    data_rows = rows[1:]

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "查询结果"

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    even_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    null_font = Font(color="999999", italic=True)

    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    data_alignment = Alignment(vertical="center", wrap_text=False)

    # 记录每列最大宽度
    col_widths = [len(h) for h in headers]

    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据行
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            # 尝试转换数字
            cell_value = value
            if value == "NULL" or value == "\\N":
                cell_value = "<NULL>"
            else:
                try:
                    if "." in value:
                        cell_value = float(value)
                    else:
                        cell_value = int(value)
                except (ValueError, TypeError):
                    pass

            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = thin_border
            cell.alignment = data_alignment

            # 交替行底色
            if row_idx % 2 == 0:
                cell.fill = even_fill

            # NULL 值标灰
            if value == "NULL" or value == "\\N":
                cell.font = null_font

            # 更新列宽
            if col_idx <= len(col_widths):
                display_len = len(str(cell_value))
                # 中文字符算 2 个宽度单位
                for ch in str(cell_value):
                    if '\u4e00' <= ch <= '\u9fff':
                        display_len += 1
                col_widths[col_idx - 1] = max(col_widths[col_idx - 1], display_len)

    # 设置列宽（加一点 padding）
    for col_idx, width in enumerate(col_widths, 1):
        adjusted_width = min(width + 4, 60)  # 最大 60
        adjusted_width = max(adjusted_width, 8)  # 最小 8
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 启用自动筛选
    if headers:
        last_col = get_column_letter(len(headers))
        last_row = len(data_rows) + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # 设置行高
    ws.row_dimensions[1].height = 24
    for row_idx in range(2, len(data_rows) + 2):
        ws.row_dimensions[row_idx].height = 20

    # 确保输出目录存在
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 保存
    wb.save(output_path)
    file_size = os.path.getsize(output_path)

    # 格式化文件大小
    if file_size < 1024:
        size_str = f"{file_size} B"
    elif file_size < 1024 * 1024:
        size_str = f"{file_size / 1024:.1f} KB"
    else:
        size_str = f"{file_size / (1024 * 1024):.1f} MB"

    print(f"[完成] Excel 导出成功")
    print(f"  文件: {output_path}")
    print(f"  大小: {size_str}")
    print(f"  数据: {len(data_rows)} 行 × {len(headers)} 列")


if __name__ == "__main__":
    main()
